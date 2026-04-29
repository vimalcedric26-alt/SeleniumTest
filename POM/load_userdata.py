# import openpyxl
#
# filename = "C:/Users/vimal/PycharmProjects/seleniumtest/data_file.xlsx"
# workbk = openpyxl.load_workbook(filename=filename)
# sheet = workbk['data_file']
# rows = sheet.max_row
# cols = sheet.max_column
# for each_row in range(1,rows+1):
#     for each_col in range(1,cols+1):
#         print(sheet.cell(row=each_row,column=each_col).value)

# Data driven testing -> Data driven testing is testing of a functionality with different data set. Application are developed first and testing begins later
# Test driven testing -> Test driven testing uses testing first and development later.

import openpyxl
class ExcelData:
    def __init__(self,filename,sheetname):
        self.fname = filename
        self.sheetname = sheetname

    def load_excel_wrkbook(self):
        work_book = openpyxl.load_workbook(filename=self.fname)
        work_sheet = work_book[self.sheetname]
        return work_book, work_sheet

    def rows_count(self,worksheet):
        rows = worksheet.max_row
        return rows

    def colmn_count(self,worksheet):
        colmn = worksheet.max_column
        return colmn

    def get_data_(self,worksheet):
        self.rows = self.rows_count(worksheet)
        self.columns = self.colmn_count(worksheet)
        user_data = []
        for each_row in range(2,self.rows+1):
            for each_col in range(1,self.columns+1):
                cell_value = worksheet.cell(row=each_row, column=each_col).value
                user_data.append(cell_value)
        return user_data

    def save_workbook(self, workbook):
        result_file = self.filename.replace(".xlsx", "_result.xlsx")
        workbook.save(result_file)

    def close_workbook(self,workbook):
        workbook.close()




# filename= "C:/Users/Kundan_Kumar/Desktop/Notes/Guvi_Notes/user_details.xlsx"
# workbk = openpyxl.load_workbook(filename=filename)
# sheet = workbk['user_data']
# rows = sheet.max_row # 5
# cols = sheet.max_column # 2
# for each_row in range(1,rows+1): # 1,6
#     for each_col in range(1,cols+1): # 1,3
#        print(sheet.cell(row=each_row,column=each_col).value) #1,1 # 1,2 # 2,1 # 2,2-----#5,1 and 5,2
#        sheet.cell(row=each_row,column=3).value = 'read'
# workbk.save(filename=filename)

