import openpyxl

workbook = openpyxl.load_workbook(filename='C:/Users/vimal/OneDrive/Desktop/Notes_Guvi/Excel_data.xlsx')
sheet = workbook["saucedemo"]
username = sheet.cell(row=1,column=1).value
password = sheet.cell(row=1,column=2).value